from __future__ import annotations
from typing import Optional
import discord
from discord.ext import commands
import logging
import os
from datetime import datetime
from zoneinfo import ZoneInfo

from config import (ADMIN_ROLE_ID, COLOR_PURPLE, COLOR_AMBER,
                    COLOR_TEAL, COLOR_GRAY, SEASON_YEAR)
from database import (get_db, config_get, get_active_season,
                      get_all_active_players, get_season_leaderboard,
                      get_week_leaderboard)
from utils.embeds import log_embed
from utils.time_utils import format_time_et

log = logging.getLogger(__name__)
ET  = ZoneInfo("America/New_York")


# ── Helpers ────────────────────────────────────────────────────────────────────

def _is_admin(interaction: discord.Interaction) -> bool:
    if not isinstance(interaction.user, discord.Member):
        return False
    return any(r.id == ADMIN_ROLE_ID for r in interaction.user.roles)


async def _log(bot: commands.Bot, description: str,
               title: str = "Stats", level: str = "info") -> None:
    ch_id = config_get("channel_logs")
    if not ch_id:
        return
    ch = bot.get_channel(int(ch_id))
    if isinstance(ch, discord.TextChannel):
        await ch.send(embed=log_embed(title, description, level))


def _latest_week(season_id: int):
    with get_db() as conn:
        return conn.execute(
            "SELECT * FROM weeks WHERE season_id=? ORDER BY week_number DESC LIMIT 1",
            (season_id,)
        ).fetchone()


# ── Streak calculation ─────────────────────────────────────────────────────────

def calculate_streaks(player_id: int) -> dict:
    """
    Returns current and longest correct pick streaks for a player,
    plus current and longest weekly win streaks.
    Picks are ordered chronologically by game kickoff time.
    """
    with get_db() as conn:
        picks = conn.execute(
            """SELECT pk.is_correct, pk.is_forfeit, g.kickoff_time
               FROM picks pk
               JOIN games g ON pk.game_id = g.id
               WHERE pk.player_id=? AND pk.is_correct IS NOT NULL
               ORDER BY g.kickoff_time ASC""",
            (player_id,)
        ).fetchall()

        weekly = conn.execute(
            """SELECT ws.weekly_rank, ws.points_earned, ws.total_possible,
                      w.week_number
               FROM weekly_scores ws
               JOIN weeks w ON ws.week_id = w.id
               WHERE ws.player_id=?
               ORDER BY w.week_number ASC""",
            (player_id,)
        ).fetchall()

    # ── Pick streak ───────────────────────────────────────────────────────────
    current_pick_streak = 0
    longest_pick_streak = 0
    running             = 0
    for p in picks:
        if p["is_forfeit"]:
            running = 0
        elif p["is_correct"]:
            running += 1
            longest_pick_streak = max(longest_pick_streak, running)
        else:
            running = 0
    current_pick_streak = running

    # ── Weekly win streak ─────────────────────────────────────────────────────
    current_week_streak = 0
    longest_week_streak = 0
    running_w           = 0
    for w in weekly:
        if w["total_possible"] == 0:
            running_w = 0
            continue
        if w["weekly_rank"] == 1:
            running_w += 1
            longest_week_streak = max(longest_week_streak, running_w)
        else:
            running_w = 0
    current_week_streak = running_w

    return {
        "current_pick_streak": current_pick_streak,
        "longest_pick_streak": longest_pick_streak,
        "current_week_streak": current_week_streak,
        "longest_week_streak": longest_week_streak,
    }


# ── Player stats embed ────────────────────────────────────────────────────────

def build_stats_embed(player: dict, season_id: int) -> discord.Embed:
    pid = player["id"]

    with get_db() as conn:
        season_totals = conn.execute(
            """SELECT
                 COALESCE(SUM(ws.points_earned),0)   AS total_points,
                 COALESCE(SUM(ws.correct_picks),0)   AS total_correct,
                 COALESCE(SUM(ws.wrong_picks),0)      AS total_wrong,
                 COALESCE(SUM(ws.forfeited_picks),0)  AS total_forfeits,
                 COALESCE(SUM(ws.total_possible),0)   AS total_possible,
                 COALESCE(MAX(ws.points_earned),0)    AS best_week,
                 COUNT(ws.week_id)                    AS weeks_played,
                 COALESCE(MIN(
                     CASE WHEN ws.total_possible > 0
                     THEN ws.points_earned END), 0)   AS worst_week
               FROM weekly_scores ws
               JOIN weeks w ON ws.week_id = w.id
               WHERE ws.player_id=? AND w.season_id=?""",
            (pid, season_id)
        ).fetchone()

        weekly_rows = conn.execute(
            """SELECT ws.points_earned, ws.correct_picks, ws.forfeited_picks,
                      ws.weekly_rank, ws.total_possible, w.week_number
               FROM weekly_scores ws
               JOIN weeks w ON ws.week_id = w.id
               WHERE ws.player_id=? AND w.season_id=?
               ORDER BY w.week_number ASC""",
            (pid, season_id)
        ).fetchall()

    streaks = calculate_streaks(pid)

    pct = (
        season_totals["total_correct"] /
        (season_totals["total_correct"] +
         season_totals["total_wrong"] +
         season_totals["total_forfeits"]) * 100
        if (season_totals["total_correct"] +
            season_totals["total_wrong"] +
            season_totals["total_forfeits"]) > 0
        else 0.0
    )

    e = discord.Embed(
        title=f"📊  {player['display_name']} — Season Stats",
        color=COLOR_PURPLE,
    )

    e.add_field(name="Total points",   value=str(season_totals["total_points"]),   inline=True)
    e.add_field(name="Correct picks",  value=str(season_totals["total_correct"]),  inline=True)
    e.add_field(name="Pick accuracy",  value=f"{pct:.1f}%",                        inline=True)
    e.add_field(name="Wrong picks",    value=str(season_totals["total_wrong"]),    inline=True)
    e.add_field(name="Missed (forfeit)",value=str(season_totals["total_forfeits"]),inline=True)
    e.add_field(name="Weeks played",   value=str(season_totals["weeks_played"]),   inline=True)
    e.add_field(name="Best week",      value=str(season_totals["best_week"]),      inline=True)
    e.add_field(name="Worst week",     value=str(season_totals["worst_week"]),     inline=True)
    e.add_field(name="\u200b",         value="\u200b",                             inline=True)

    # Streaks
    e.add_field(
        name="Pick streaks",
        value=(
            f"Current: **{streaks['current_pick_streak']}** correct in a row\n"
            f"Longest: **{streaks['longest_pick_streak']}**"
        ),
        inline=True,
    )
    e.add_field(
        name="Weekly win streaks",
        value=(
            f"Current: **{streaks['current_week_streak']}** week wins in a row\n"
            f"Longest: **{streaks['longest_week_streak']}**"
        ),
        inline=True,
    )
    e.add_field(name="\u200b", value="\u200b", inline=True)

    # Weekly breakdown table
    if weekly_rows:
        week_lines = []
        for w in weekly_rows:
            rank_str = f"#{w['weekly_rank']}" if w["weekly_rank"] else "—"
            forfeit  = f" · {w['forfeited_picks']}M" if w["forfeited_picks"] else ""
            week_lines.append(
                f"Wk {w['week_number']:>2}  "
                f"{str(w['points_earned']):>4} pts  "
                f"{str(w['correct_picks']):>2}/{str(w['total_possible']):>3} pts possible  "
                f"{rank_str}{forfeit}"
            )
        e.add_field(
            name="Week-by-week",
            value="```" + "\n".join(week_lines) + "```",
            inline=False,
        )

    e.set_footer(text=f"M = missed (forfeit)  ·  {SEASON_YEAR} season")
    return e


# ── Pick history embed ────────────────────────────────────────────────────────

def build_history_embed(player: dict, week_number: int,
                        week_id: int) -> discord.Embed:
    with get_db() as conn:
        picks = conn.execute(
            """SELECT pk.picked_team, pk.confidence_points,
                      pk.is_correct, pk.is_forfeit,
                      g.home_team, g.away_team, g.home_rank, g.away_rank,
                      g.winner, g.kickoff_time, g.channel
               FROM picks pk
               JOIN games g ON pk.game_id = g.id
               WHERE pk.player_id=? AND g.week_id=?
               ORDER BY pk.confidence_points DESC""",
            (player["id"], week_id)
        ).fetchall()

        score = conn.execute(
            "SELECT * FROM weekly_scores WHERE player_id=? AND week_id=?",
            (player["id"], week_id)
        ).fetchone()

    e = discord.Embed(
        title=f"📋  {player['display_name']} — Week {week_number} picks",
        color=COLOR_TEAL,
    )

    if not picks:
        e.description = "No picks found for this week."
        return e

    lines = []
    for p in picks:
        if p["is_forfeit"]:
            icon = "❌"
            desc = (
                f"MISSED — {p['home_team']} vs {p['away_team']} "
                f"*(winner: {p['winner'] or 'TBD'})*"
            )
        elif p["is_correct"] is None:
            icon = "⏳"
            desc = f"{p['picked_team']}  ({p['home_team']} vs {p['away_team']})"
        elif p["is_correct"]:
            icon = "✅"
            desc = f"{p['picked_team']}  ({p['home_team']} vs {p['away_team']})"
        else:
            icon = "❌"
            desc = (
                f"{p['picked_team']}  ({p['home_team']} vs {p['away_team']}) "
                f"*(winner: {p['winner']})*"
            )

        lines.append(
            f"`{p['confidence_points']:>2}` {icon} {desc}"
        )

    e.description = "\n".join(lines)

    if score:
        pct = (
            score["correct_picks"] / len(picks) * 100
            if picks else 0
        )
        summary = (
            f"**{score['points_earned']} pts**  ·  "
            f"{score['correct_picks']} correct  ·  "
            f"{score['wrong_picks']} wrong"
        )
        if score["forfeited_picks"]:
            summary += f"  ·  {score['forfeited_picks']} missed"
        if score["weekly_rank"]:
            summary += f"  ·  Rank #{score['weekly_rank']}"
        e.add_field(name="Summary", value=summary, inline=False)

    return e


# ── Week history select ────────────────────────────────────────────────────────

class WeekHistorySelect(discord.ui.Select):
    def __init__(self, player: dict, weeks: list):
        self.player = player
        options = [
            discord.SelectOption(
                label=f"Week {w['week_number']}",
                value=f"{w['id']}:{w['week_number']}",
            )
            for w in weeks[:25]
        ]
        super().__init__(placeholder="Select a week…", options=options)

    async def callback(self, interaction: discord.Interaction) -> None:
        week_id, week_number = self.values[0].split(":")
        embed = build_history_embed(
            self.player, int(week_number), int(week_id)
        )
        await interaction.response.edit_message(embed=embed, view=self.view)


class PickHistoryView(discord.ui.View):
    def __init__(self, player: dict, weeks: list):
        super().__init__(timeout=120)
        self.add_item(WeekHistorySelect(player, weeks))

        close = discord.ui.Button(
            label="Close", style=discord.ButtonStyle.secondary
        )
        close.callback = self._close
        self.add_item(close)

    async def _close(self, interaction: discord.Interaction) -> None:
        await interaction.response.edit_message(
            content="Closed.", embed=None, view=None
        )
        self.stop()


# ── Excel export ──────────────────────────────────────────────────────────────

def _export_week_xlsx(week_id: int, week_number: int,
                      season_year: int) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    with get_db() as conn:
        players = conn.execute(
            "SELECT * FROM players WHERE status IN ('active','withdrawn') "
            "ORDER BY display_name"
        ).fetchall()
        games = conn.execute(
            "SELECT * FROM games WHERE week_id=? ORDER BY kickoff_time",
            (week_id,)
        ).fetchall()
        scores = conn.execute(
            "SELECT * FROM weekly_scores WHERE week_id=?", (week_id,)
        ).fetchall()
        all_picks = conn.execute(
            """SELECT pk.*, g.home_team, g.away_team
               FROM picks pk
               JOIN games g ON pk.game_id = g.id
               WHERE g.week_id=?""",
            (week_id,)
        ).fetchall()

    score_map = {s["player_id"]: dict(s) for s in scores}
    picks_map: dict = {}
    for p in all_picks:
        picks_map.setdefault(p["player_id"], {})[p["game_id"]] = dict(p)

    wb = Workbook()

    # ── Scores sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = f"Week {week_number} Scores"

    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    bold        = Font(bold=True, name="Arial")
    normal      = Font(name="Arial")

    headers = ["Player", "Points", "Correct", "Wrong",
               "Missed", "Rank", "Total Possible"]
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = Alignment(horizontal="center")

    for row_idx, player in enumerate(players, 2):
        s = score_map.get(player["id"], {})
        ws.cell(row=row_idx, column=1, value=player["display_name"]).font = bold
        ws.cell(row=row_idx, column=2, value=s.get("points_earned",  0)).font = normal
        ws.cell(row=row_idx, column=3, value=s.get("correct_picks",  0)).font = normal
        ws.cell(row=row_idx, column=4, value=s.get("wrong_picks",    0)).font = normal
        ws.cell(row=row_idx, column=5, value=s.get("forfeited_picks",0)).font = normal
        ws.cell(row=row_idx, column=6, value=s.get("weekly_rank",   "")).font = normal
        ws.cell(row=row_idx, column=7, value=s.get("total_possible", 0)).font = normal

    for col in range(1, 8):
        ws.column_dimensions[
            ws.cell(row=1, column=col).column_letter
        ].width = 16

    # ── Picks sheet ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(f"Week {week_number} Picks")

    pick_headers = (
        ["Player"] +
        [f"{g['home_team']} vs {g['away_team']}" for g in games] +
        ["Points", "Correct"]
    )
    for col, h in enumerate(pick_headers, 1):
        cell            = ws2.cell(row=1, column=col, value=h)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = Alignment(horizontal="center", wrap_text=True)
        ws2.column_dimensions[cell.column_letter].width = 20

    for row_idx, player in enumerate(players, 2):
        ws2.cell(row=row_idx, column=1,
                 value=player["display_name"]).font = bold
        player_picks = picks_map.get(player["id"], {})
        for col_idx, game in enumerate(games, 2):
            pick = player_picks.get(game["id"])
            if pick:
                if pick["is_forfeit"]:
                    val = "MISSED"
                else:
                    correct = pick.get("is_correct")
                    val = (
                        f"✓ {pick['picked_team']}" if correct
                        else f"✗ {pick['picked_team']}" if correct is not None
                        else pick["picked_team"] or ""
                    )
                    val = f"{val} ({pick['confidence_points']}pts)"
            else:
                val = "—"
            ws2.cell(row=row_idx, column=col_idx, value=val).font = normal

        s = score_map.get(player["id"], {})
        ws2.cell(row=row_idx, column=len(games)+2,
                 value=s.get("points_earned", 0)).font = normal
        ws2.cell(row=row_idx, column=len(games)+3,
                 value=s.get("correct_picks", 0)).font = normal

    path = os.path.join("data", f"Week_{week_number}_Export.xlsx")
    wb.save(path)
    return path


def _export_season_xlsx(season_id: int, season_year: int) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    season_rows = get_season_leaderboard(season_id)

    with get_db() as conn:
        weeks = conn.execute(
            "SELECT * FROM weeks WHERE season_id=? ORDER BY week_number",
            (season_id,)
        ).fetchall()
        players = conn.execute(
            "SELECT * FROM players WHERE status IN ('active','withdrawn') "
            "ORDER BY display_name"
        ).fetchall()

    wb = Workbook()

    # ── Season standings sheet ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Season Standings"

    header_fill = PatternFill("solid", fgColor="BA7517")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    bold        = Font(bold=True, name="Arial")
    normal      = Font(name="Arial")

    headers = ["Rank", "Player", "Total Points", "Correct",
               "Wrong", "Missed", "% Correct", "Best Week", "Worst Week"]
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 14

    for row_idx, r in enumerate(season_rows, 2):
        pct = (r["total_correct"] / r["total_possible"] * 100
               if r["total_possible"] else 0)
        ws.cell(row=row_idx, column=1, value=row_idx - 1).font   = normal
        ws.cell(row=row_idx, column=2, value=r["display_name"]).font = bold
        ws.cell(row=row_idx, column=3, value=r["total_points"]).font  = normal
        ws.cell(row=row_idx, column=4, value=r["total_correct"]).font = normal
        ws.cell(row=row_idx, column=5, value=r["total_wrong"]).font   = normal
        ws.cell(row=row_idx, column=6, value=r["total_forfeits"]).font= normal
        ws.cell(row=row_idx, column=7,
                value=round(pct, 1)).number_format = "0.0%"
        ws.cell(row=row_idx, column=7).font = normal
        ws.cell(row=row_idx, column=8, value=r["best_week"]).font  = normal
        ws.cell(row=row_idx, column=9, value=r["worst_week"]).font = normal

    # ── Weekly breakdown sheet ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Weekly Breakdown")

    week_headers = ["Player"] + [f"Week {w['week_number']}" for w in weeks] + ["Total"]
    for col, h in enumerate(week_headers, 1):
        cell      = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions[cell.column_letter].width = 12

    with get_db() as conn:
        all_weekly = conn.execute(
            """SELECT ws.player_id, ws.points_earned, w.week_number
               FROM weekly_scores ws
               JOIN weeks w ON ws.week_id = w.id
               WHERE w.season_id=?""",
            (season_id,)
        ).fetchall()

    wk_map: dict = {}
    for row in all_weekly:
        wk_map.setdefault(row["player_id"], {})[row["week_number"]] = row["points_earned"]

    for row_idx, player in enumerate(players, 2):
        ws2.cell(row=row_idx, column=1, value=player["display_name"]).font = bold
        total = 0
        for col_idx, w in enumerate(weeks, 2):
            pts = wk_map.get(player["id"], {}).get(w["week_number"], 0)
            ws2.cell(row=row_idx, column=col_idx, value=pts).font = normal
            total += pts
        ws2.cell(row=row_idx, column=len(weeks)+2, value=total).font = bold

    path = os.path.join("data", f"{season_year}_Season_Export.xlsx")
    wb.save(path)
    return path


# ── Stats cog and wiring ──────────────────────────────────────────────────────

class StatsCog(commands.Cog):
    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot


async def open_my_stats(interaction: discord.Interaction) -> None:
    """Called from the picks hub 'My stats' button."""
    from database import get_player_by_discord_id
    player = get_player_by_discord_id(str(interaction.user.id))
    if not player:
        await interaction.response.send_message(
            "You're not registered yet.", ephemeral=True
        )
        return
    if player["status"] != "active":
        await interaction.response.send_message(
            "Stats are only available for active players.", ephemeral=True
        )
        return

    season = get_active_season()
    if not season:
        await interaction.response.send_message(
            "No active season.", ephemeral=True
        )
        return

    embed = build_stats_embed(dict(player), season["id"])
    await interaction.response.send_message(embed=embed, ephemeral=True)


async def open_pick_history(interaction: discord.Interaction) -> None:
    """Called from the picks hub 'My picks' button for history view."""
    from database import get_player_by_discord_id
    player = get_player_by_discord_id(str(interaction.user.id))
    if not player:
        await interaction.response.send_message(
            "You're not registered yet.", ephemeral=True
        )
        return

    season = get_active_season()
    if not season:
        await interaction.response.send_message(
            "No active season.", ephemeral=True
        )
        return

    with get_db() as conn:
        weeks = conn.execute(
            """SELECT w.id, w.week_number FROM weeks w
               WHERE w.season_id=?
               ORDER BY w.week_number DESC""",
            (season["id"],)
        ).fetchall()

    if not weeks:
        await interaction.response.send_message(
            "No weeks played yet.", ephemeral=True
        )
        return

    # Show most recent week by default
    first_week = weeks[0]
    embed = build_history_embed(
        dict(player), first_week["week_number"], first_week["id"]
    )
    view = PickHistoryView(dict(player), [dict(w) for w in weeks])
    await interaction.response.send_message(
        embed=embed, view=view, ephemeral=True
    )


async def export_week(interaction: discord.Interaction,
                      bot: commands.Bot) -> None:
    """Admin export for current week."""
    await interaction.response.defer(ephemeral=True)

    season = get_active_season()
    if not season:
        await interaction.followup.send("No active season.", ephemeral=True)
        return

    with get_db() as conn:
        week = conn.execute(
            "SELECT * FROM weeks WHERE season_id=? ORDER BY week_number DESC LIMIT 1",
            (season["id"],)
        ).fetchone()

    if not week:
        await interaction.followup.send("No week loaded.", ephemeral=True)
        return

    try:
        path = _export_week_xlsx(
            week["id"], week["week_number"], season["year"]
        )
    except Exception as exc:
        await interaction.followup.send(
            f"Export failed: {exc}", ephemeral=True
        )
        return

    ch_id = config_get("channel_logs")
    if ch_id:
        ch = bot.get_channel(int(ch_id))
        if isinstance(ch, discord.TextChannel):
            await ch.send(
                content=f"Week {week['week_number']} export generated.",
                file=discord.File(path),
            )

    await interaction.followup.send(
        f"✅ Week {week['week_number']} export posted to #cfcp-logs.",
        ephemeral=True,
    )


async def export_season(bot: commands.Bot) -> None:
    """Generate and post the full season export — called at season end."""
    season = get_active_season()
    if not season:
        return

    try:
        path = _export_season_xlsx(season["id"], season["year"])
    except Exception as exc:
        log.error(f"Season export failed: {exc}", exc_info=True)
        return

    ch_id = config_get("channel_logs")
    if ch_id:
        ch = bot.get_channel(int(ch_id))
        if isinstance(ch, discord.TextChannel):
            await ch.send(
                content=f"📊 {season['year']} season export — final standings.",
                file=discord.File(path),
            )


async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(StatsCog(bot))
    log.info("StatsCog loaded.")
